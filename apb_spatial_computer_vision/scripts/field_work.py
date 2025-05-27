import os,sys
from openpyxl import load_workbook

# Get the root directory
root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Add the root directory to sys.path
sys.path.append(root_dir)

from apb_spatial_computer_vision import *

import geopandas as gpd , pandas as pd, numpy as np
from pyproj import Proj,Transformer,CRS

np.set_printoptions(suppress=True)
# Custom formatting function
def sci_notation_formatter(x):
    if abs(x) < 1e-6:
        return f"{x:.2e}"
    else:
        return f"{x}"


np.set_printoptions(suppress=True)

def vector_deg_to_DMS(alpha):
        alpha=np.array(alpha)
        dd=np.abs(np.astype(alpha,np.uint8))
        min=np.abs(np.astype((np.abs(alpha)-dd)*60,np.uint8))
        sec=np.abs(((((np.abs(alpha)-dd)*60)-min)*60))
        signs=np.array(np.sign(alpha))
        return ((signs*dd).astype(np.int8),(signs*min).astype(np.int8),signs*sec)
        
def quality_control_point_creation(from_crs=25831,geographic_crs=4258):
    """Data processing from the points with averaging
    """
    x_col='Coordenada X'
    y_col='Coordenada Y'
    z_col='Altura Ortom.'
    h_col='Altura Elipsoidal'
    id_col='Punto Id'
    std_x='D.Est X'
    std_y='D.Est Y'
    std_H='D.Est Altura Ortom'
    std_h='D.Est. Alt. Elip.'
    N='Ondulación Geoidal'
    date='Fecha/Hora'
    
    df=pd.read_csv(os.path.join(DATA_DIR,'INFINITY_COMPROV_ORTO.txt'),header=0,sep=';')
    gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df[x_col],df[y_col]),crs=from_crs)
    
    projection=Proj(from_crs)
    transformer=Transformer.from_crs(from_crs,geographic_crs)

    
    gdf=gdf.to_crs(geographic_crs)
    gdf['LAT'],gdf['LON'],gdf['h']=transformer.transform(gdf[x_col],gdf[y_col],gdf[z_col])
    gdf['LAT_moved'],gdf['LON_moved'],gdf['h_moved']=transformer.transform(gdf[x_col]+gdf[std_x],gdf[y_col]+gdf[std_y],gdf[z_col])
    gdf['STD_LAT']=np.array([f"{x:.6f}" for x in (np.array(np.abs(gdf['LAT_moved']-gdf['LAT'])*3600)).round(6)])
    gdf['STD_LON']=np.array([f"{x:.6f}" for x in (np.array(np.abs(gdf['LON_moved']-gdf['LON'])*3600)).round(6)])

    ltd,ltm,lts=vector_deg_to_DMS(gdf['LAT'])
    gdf[['LTD','LTM','LTS']]=pd.DataFrame({'LTD':ltd,'LTM':ltm,'LTS':lts.round(3)})
    
    lnd,lnm,lns=vector_deg_to_DMS(gdf['LON'])
    gdf[['LND','LNM','LNS']]=pd.DataFrame({'LND':lnd,'LNM':lnm,'LNS':lns.round(3)})
    
    gdf['K']=projection.get_factors(gdf['LON'],gdf['LAT']).meridional_scale.round(7)
    gdf['DATA']=np.array([i.split(' ')[0] for i in gdf[date]]).astype(str)
    file_path = os.path.join(DATA_DIR,'SURVEYING.xlsx')
    
    # # WAIT FOR GTX FILE
    # geoid_tiff=os.path.join(DATA_DIR,'geodesy','EGM08D595_4258.tif')
    # "gdal_translate c:\\dev\\TFG\\data\\geodesy\\EGM08D595_4258.tif c:\\dev\\TFG\\data\\geodesy\\EGM08D595_4258.gtx"
    # geoid_grid=os.path.join(DATA_DIR,'geodesy','EGM08D595_4258.gtx')
    # crs_egm08d595 = CRS.from_proj4(
    #      f'''+proj=vgridshift +grids={geoid_grid}''')
    
    aux_in=os.path.join(BASE_DIR,'input.dat')
    aux_out=os.path.join(BASE_DIR,'output_UTM_eh.dat')
    
    
    with open(aux_in, "w") as f:
        f.write(df[[x_col,y_col,z_col]].to_string(header=0,justify="right"))
    f.close()
    import subprocess
    result = subprocess.run(os.path.join(SHELL_DIR,'UTMOHehx.exe'), shell=True, capture_output=True, text=True)
    print(result.stdout)
    prev_df=pd.read_csv(aux_out,header=0,sep='\s+',engine='python',usecols=lambda x: x != '#').drop(index=0)
    os.remove(aux_in)
    os.remove(aux_out)
    
    valid_columns=prev_df.columns[1:] 
    prev_df=prev_df.drop(columns=prev_df.columns[-1])
    prev_df.columns=valid_columns
    prev_df=prev_df.reset_index(drop=True)
    gdf['N']=prev_df['Ondulacio']
    gdf['h']=prev_df['eh']
    print(prev_df)
    
    workbook = load_workbook(file_path,data_only=False)
    # for sheet in workbook.sheetnames:
    #     workbook[sheet].sheet_state = 'visible'
    # sheet = workbook.active
    for sheet in workbook.sheetnames:
        workbook[sheet].sheet_state = 'visible'

    # Select the active sheet
    sheet = workbook.active

    df_excel = pd.read_excel(file_path, sheet_name=sheet.title)
    df_excel = df_excel.reindex(range(len(gdf)), fill_value=np.nan)
    df_excel[['X_UTM','Y_UTM','LAT','LTD','LTM','LTS','LON','LND','LNM','LNS','K','H','h','N','CODI','STD_X','STD_Y','STD_H','STD_h','STD_LAT','STD_LON','DATA']] = gdf[[x_col,y_col,'LAT','LTD','LTM','LTS','LON','LND','LNM','LNS','K',z_col,h_col,'N',id_col,std_x,std_y,std_H,std_h,'STD_LAT','STD_LON','DATA']].values
    df_excel['CRS']='ETRS89'
    df_excel['GEOIDE']='EGM08D595'
    df[[x_col,y_col,z_col,std_x,std_y]].round(3)
    df_excel=df_excel.astype(str)
    df_excel.to_excel(file_path,engine='openpyxl', sheet_name=sheet.title, index=False)
    #workbook.save(file_path)
    workbook.close()
    df_excel.to_csv(os.path.join(OUT_DIR,'ORTOFOTO_CHECK.CSV'))
    gdf.to_file(os.path.join(DATA_DIR,'ORTOFOTO_CHECK.geojson'))
quality_control_point_creation()