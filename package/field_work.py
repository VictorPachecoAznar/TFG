import os,sys
from openpyxl import load_workbook

# Get the root directory
root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Add the root directory to sys.path
sys.path.append(root_dir)

from package import *

import geopandas as gpd , pandas as pd, numpy as np
from pyproj import Proj,Transformer,CRS

def quality_control_point_creation(from_crs=25831,geographic_crs=4258):
    """Data processing from the points with averaging
    """
    df=pd.read_csv(os.path.join(DATA_DIR,'COMPROV ORTO.txt'),sep=' ')
    gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df['X'],df['Y']),crs=from_crs)
    
    projection=Proj(from_crs)
    transformer=Transformer.from_crs(from_crs,geographic_crs)

    
    gdf=gdf.to_crs(geographic_crs)
    gdf['LAT'],gdf['LON'],gdf['h']=transformer.transform(gdf['X'],gdf['Y'],gdf['Z'])
    gdf['K']=projection.get_factors(gdf['LON'],gdf['LAT']).meridional_scale
    
    file_path = os.path.join(DATA_DIR,'SURVEYING.xlsx')
    
    # WAIT FOR GTX FILE
    geoid_tiff=os.path.join(DATA_DIR,'geodesy','EGM08D595_4258.tif')
    "gdal_translate c:\\dev\\TFG\\data\\geodesy\\EGM08D595_4258.tif c:\\dev\\TFG\\data\\geodesy\\EGM08D595_4258.gtx"
    geoid_grid=os.path.join(DATA_DIR,'geodesy','EGM08D595_4258.gtx')
    crs_egm08d595 = CRS.from_proj4(
         f'''+proj=vgridshift +grids={geoid_grid}''')
    
    aux_in=os.path.join(BASE_DIR,'input.dat')
    aux_out=os.path.join(BASE_DIR,'output_UTM_eh.dat')
    
    
    with open(aux_in, "w") as f:
        f.write(df[['X','Y','Z']].to_string(header=None,justify="right"))
    f.close()
    import subprocess
    result = subprocess.run(os.path.join(SHELL_DIR,'UTMOHehx.exe'), shell=True, capture_output=True, text=True)
    print(result.stdout)
    prev_df=pd.read_csv(aux_out,header=0,sep='\s+',engine='python',usecols=lambda x: x != '#').drop(index=0)
    os.remove(aux_in)
    os.remove(aux_out)
    
    valid_columns=prev_df.columns[1:] 
    prev_df=prev_df.drop(columns=prev_df.columns[-1])
    prev_df.columns=valid_columns
    gdf['h']=prev_df['eh']
    
    workbook = load_workbook(file_path,data_only=False)
    # for sheet in workbook.sheetnames:
    #     workbook[sheet].sheet_state = 'visible'
    # sheet = workbook.active
    for sheet in workbook.sheetnames:
        workbook[sheet].sheet_state = 'visible'

    # Select the active sheet
    sheet = workbook.active

    df_excel = pd.read_excel(file_path, sheet_name=sheet.title)
    df_excel = df_excel.reindex(range(len(gdf)), fill_value=np.nan)
    df_excel[['X_UTM','Y_UTM','LAT','LON','K','H','h','CODI']] = gdf[['X','Y','LAT','LON','K','Z','h','P']].values
    df_excel['CRS']='ETRS89'
    df_excel['GEOIDE']='EGM08D595'
    
    df_excel.to_excel(file_path,engine='openpyxl', sheet_name=sheet.title, index=False)
    #workbook.save(file_path)
    workbook.close()
    df_excel.to_csv(os.path.join(OUT_DIR,'ORTOFOTO_CHECK.CSV'))
    gdf.to_file(os.path.join(DATA_DIR,'ORTOFOTO_CHECK.geojson'))
quality_control_point_creation()